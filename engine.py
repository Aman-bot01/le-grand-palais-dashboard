"""
PFH / EdgeLabs Player Segmentation - data + clustering engine
=============================================================
SAFE-CLONE VERSION: connects to a local, 100% synthetic DuckDB file
(synthetic.duckdb, built by synthetic_data.py) instead of the real SQL Server /
Oracle warehouse. No network connection, no real credentials, no real data.

Kept the same public function signatures the callers (launch.py,
launch_dashboard_v2.py) use -- get_connection(), query_df(), and
load_game_catalog_with_fallback() -- so those files needed minimal changes
beyond swapping T-SQL-only syntax for its DuckDB equivalent (see the
DATEDIFF/DATEADD/DATEPART/[bracket] notes in launch.py and launch_dashboard_v2.py).
"""
from __future__ import annotations

import os
import datetime as dt

import numpy as np
import pandas as pd
import duckdb

from sklearn.preprocessing import StandardScaler
from sklearn.cluster import KMeans, MiniBatchKMeans

HERE = os.path.dirname(os.path.abspath(__file__))
CACHE_DIR = os.path.join(HERE, "cache")
DATA_DIR = os.path.join(HERE, "data")
os.makedirs(CACHE_DIR, exist_ok=True)

DB_PATH = os.path.join(HERE, "synthetic.duckdb")

WINDOW_START = os.getenv("WINDOW_START", "2024-09-01")
EL_LOOKBACK_DAYS = int(os.getenv("EL_LOOKBACK_DAYS", "30"))

# Platform modes -----------------------------------------------------------
MODES = {
    "PFH": dict(platform="Pong", casinos=["PFH"], dim="location",
                unit_label="State / Store"),
    "EdgeLabs": dict(platform="EdgeLabs", casinos=None, dim="casino",
                     unit_label="Casino"),
}
GAME_PREFIXES = ("V1", "V2")

HANDLE_SQL = "CAST(TotalBet AS FLOAT)/100.0"
WIN_SQL = "CAST(TotalWin AS FLOAT)/100.0"

FEATURE_COLS = [
    "log_LCI", "log_CI30", "recency_score", "recent_active_pct",
    "log_LAD", "log_avg_bet", "log_spins_day", "trend_clipped",
]
CLUSTER_ORDER = ["VIP", "High-Roller", "Steady Regular",
                 "Surging Spender", "Lapsed", "Light Tried"]
CLUSTER_COLORS = {
    "VIP": "#2563eb", "High-Roller": "#7c3aed", "Steady Regular": "#0d9488",
    "Surging Spender": "#f59e0b", "Lapsed": "#ef4444", "Light Tried": "#94a3b8",
}


# ── Connection ───────────────────────────────────────────────────────
def get_connection(query_timeout: int = 900):
    """Return a DuckDB connection to the local synthetic database, with a
    handful of SQL-Server-compatibility macros registered so the (unmodified
    where possible) T-SQL query strings in launch.py / launch_dashboard_v2.py
    run against it without a full rewrite."""
    if not os.path.exists(DB_PATH):
        # First run on a fresh checkout/deployment (e.g. Streamlit Community Cloud) --
        # build the synthetic database on the spot instead of failing, so the app works
        # from a bare `git clone` + `streamlit run` with no manual setup step. Seeded
        # (SEED=42 in synthetic_data.py), so this is fully deterministic.
        import synthetic_data
        synthetic_data.main()
    con = duckdb.connect(DB_PATH, read_only=False)

    # GETDATE() / ISNULL() -- direct analogues, no argument-shape surprises.
    con.execute("CREATE OR REPLACE MACRO GETDATE() AS now();")
    # ISNULL is quoted here because DuckDB's parser reserves the bareword ISNULL
    # token (it's part of "IS NULL" tokenization); the quoted CREATE still lets
    # every unquoted ISNULL(...) call site in the calling SQL resolve normally.
    con.execute('CREATE OR REPLACE MACRO "ISNULL"(a, b) AS COALESCE(a, b);')

    # DATEDIFF/DATEADD/DATEPART -- T-SQL takes an unquoted unit keyword
    # (day/week/month/year); DuckDB's date_diff()/date_part() take it as a
    # string, and there is no DATEADD equivalent at all. The calling files
    # have been text-rewritten to quote the unit (DATEDIFF(day, -> DATEDIFF('day',
    # etc.) so these macros can dispatch on it as a string literal.
    con.execute("""
        CREATE OR REPLACE MACRO DATEDIFF(part, startdate, enddate) AS
            date_diff(part, CAST(startdate AS TIMESTAMP), CAST(enddate AS TIMESTAMP));
    """)
    con.execute("""
        CREATE OR REPLACE MACRO DATEADD(part, num, dt) AS
            CASE part
                WHEN 'day'   THEN CAST(dt AS TIMESTAMP) + (num * INTERVAL 1 DAY)
                WHEN 'week'  THEN CAST(dt AS TIMESTAMP) + (num * INTERVAL 1 WEEK)
                WHEN 'month' THEN CAST(dt AS TIMESTAMP) + (num * INTERVAL 1 MONTH)
                WHEN 'year'  THEN CAST(dt AS TIMESTAMP) + (num * INTERVAL 1 YEAR)
            END;
    """)
    con.execute("""
        CREATE OR REPLACE MACRO DATEPART(part, dt) AS
            date_part(part, CAST(dt AS TIMESTAMP));
    """)
    return con


def query_df(conn, sql: str) -> pd.DataFrame:
    return conn.execute(sql).fetchdf()


def _esc(s: str) -> str:
    return str(s).replace("'", "''")


def _scope_where(mode: str) -> str:
    m = MODES[mode]
    w = f"PlatformName='{m['platform']}'"
    if m["casinos"]:
        w += " AND CasinoName IN (" + ",".join("'" + c + "'" for c in m["casinos"]) + ")"
    return w


def default_window(mode: str, as_of: dt.date):
    if mode == "PFH":
        return dt.date.fromisoformat(WINDOW_START), as_of
    return as_of - dt.timedelta(days=EL_LOOKBACK_DAYS), as_of


# ── Raw pulls ────────────────────────────────────────────────────────
def get_latest_date(conn, mode: str) -> dt.date:
    df = query_df(conn, f"""SELECT MAX("Date") d FROM BetSpinSummaryCashView3
        WHERE {_scope_where(mode)}""")
    return pd.to_datetime(df["d"].iloc[0]).date()


def get_window_start(conn, mode: str) -> dt.date:
    try:
        df = query_df(conn, f"""SELECT MIN("Date") d FROM BetSpinSummaryCashView3
            WHERE {_scope_where(mode)}""")
        detected = pd.to_datetime(df["d"].iloc[0]).date()
        floor = dt.date.fromisoformat(WINDOW_START)
        return max(detected, floor)
    except Exception:
        return dt.date.fromisoformat(WINDOW_START)


def load_daily_aggregate(conn, start, end, mode: str) -> pd.DataFrame:
    unit = "CAST(StoreNumber AS VARCHAR(50))" if MODES[mode]["dim"] == "location" else "CasinoName"
    sql = f"""
        SELECT AccountNumber, {unit} AS UnitKey, AggregatorName AS Aggregator, "Date",
               SUM({HANDLE_SQL}) AS Handle, SUM({WIN_SQL}) AS Win, SUM(Spins) AS Spins
        FROM BetSpinSummaryCashView3
        WHERE {_scope_where(mode)} AND "Date" >= '{start}' AND "Date" <= '{end}'
        GROUP BY AccountNumber, {unit}, AggregatorName, "Date" """
    df = query_df(conn, sql)
    cols = ["AccountNumber", "UnitKey", "Aggregator", "Date", "Handle", "Win", "Spins"]
    if df.empty:
        return pd.DataFrame(columns=cols)
    df["Date"] = pd.to_datetime(df["Date"])
    df["Handle"] = pd.to_numeric(df["Handle"]).astype(float)
    df["Win"] = pd.to_numeric(df["Win"]).astype(float)
    df["Spins"] = pd.to_numeric(df["Spins"]).astype("int64")
    df["AccountNumber"] = df["AccountNumber"].astype(str)
    df["UnitKey"] = df["UnitKey"].astype(str).str.strip()
    df["Aggregator"] = df["Aggregator"].astype(str).str.strip()
    return df[cols]


def load_locations(conn) -> pd.DataFrame:
    df = query_df(conn, """SELECT LocationId, StateProv, City, BusinessName,
                                  Distributor, Operator, PFHEnabled
                           FROM CrmLocationView""")
    df["LocationId"] = df["LocationId"].astype(str).str.strip()
    return df


def get_active_location_count(conn, game_id, mode: str, as_of_days: int = 14) -> dict:
    gid = str(game_id).strip()
    cutoff = f"DATEADD('day', -{int(as_of_days)}, CAST(GETDATE() AS DATE))"

    if mode in ("V2", "V1"):
        sql = f"""
        SELECT
            CAST(b.StoreNumber AS VARCHAR(50)) AS loc,
            MAX(loc.StateProv) AS state_prov
        FROM TaskHandlerBetSpinSummary b
        LEFT JOIN CrmLocationView loc
            ON CAST(b.StoreNumber AS VARCHAR(50)) = CAST(loc.LocationId AS VARCHAR(50))
        WHERE b.CasinoName IN ('vendor1', 'vendor2')
          AND b.GameId = {int(gid)}
          AND CAST(b.SummaryDate AS DATE) >= {cutoff}
          AND b.Spins > 0
        GROUP BY CAST(b.StoreNumber AS VARCHAR(50))
        """
        df = query_df(conn, sql)
        if df.empty:
            return {"count": 0, "unit_label": "locations", "by_state": {}}
        by_state: dict[str, int] = {}
        for _, row in df.iterrows():
            st = str(row.get("state_prov") or "Unknown").strip()
            by_state[st] = by_state.get(st, 0) + 1
        return {"count": len(df), "unit_label": "locations", "by_state": by_state}
    elif MODES.get(mode, {}).get("dim") == "location":
        sql = f"""
        SELECT
            CAST(b.StoreNumber AS VARCHAR(50)) AS loc,
            MAX(loc.StateProv) AS state_prov
        FROM BetSpinSummaryCashView3 b
        LEFT JOIN CrmLocationView loc
            ON CAST(b.StoreNumber AS VARCHAR(50)) = CAST(loc.LocationId AS VARCHAR(50))
        WHERE {_scope_where(mode)}
          AND b.GameId = '{_esc(gid)}'
          AND CAST(b."Date" AS DATE) >= {cutoff}
          AND b.Spins > 0
        GROUP BY CAST(b.StoreNumber AS VARCHAR(50))
        """
        df = query_df(conn, sql)
        if df.empty:
            return {"count": 0, "unit_label": "locations", "by_state": {}}
        by_state: dict[str, int] = {}
        for _, row in df.iterrows():
            st = str(row.get("state_prov") or "Unknown").strip()
            by_state[st] = by_state.get(st, 0) + 1
        return {"count": len(df), "unit_label": "locations", "by_state": by_state}
    else:
        _el_scope = _scope_where(mode) if mode in MODES else "PlatformName='EdgeLabs'"
        sql = f"""
        SELECT COUNT(DISTINCT CasinoName) AS cnt
        FROM BetSpinSummaryCashView3
        WHERE {_el_scope}
          AND TRY_CAST(GameId AS INT) = {int(gid)}
          AND CAST("Date" AS DATE) >= {cutoff}
          AND Spins > 0
        """
        df = query_df(conn, sql)
        cnt = int(df["cnt"].iloc[0]) if not df.empty and "cnt" in df.columns else 0
        return {"count": cnt, "unit_label": "casinos", "by_state": {}}


# ── Incremental cache (parquet, per mode; extends both directions) ───
def _cache_path(mode: str) -> str:
    return os.path.join(CACHE_DIR, f"daily_{mode}_v2.parquet")


def _slice(df, start, end):
    s, e = pd.to_datetime(start), pd.to_datetime(end)
    return df[(df["Date"] >= s) & (df["Date"] <= e)].copy()


def refresh_daily_cache(conn, mode: str, start, end, full: bool = False) -> pd.DataFrame:
    path = _cache_path(mode)
    cached = None
    if not full and os.path.exists(path):
        try:
            cached = pd.read_parquet(path)
            cached["Date"] = pd.to_datetime(cached["Date"])
        except Exception:
            cached = None

    if cached is None or cached.empty:
        fresh = load_daily_aggregate(conn, start, end, mode)
        fresh.to_parquet(path, index=False)
        return _slice(fresh, start, end)

    parts = [cached]
    cmin, cmax = cached["Date"].min().date(), cached["Date"].max().date()
    req_start = pd.to_datetime(start).date()
    req_end = pd.to_datetime(end).date()
    if req_start < cmin:
        parts.append(load_daily_aggregate(conn, start, cmin - dt.timedelta(days=1), mode))
    if req_end > cmax:
        parts.append(load_daily_aggregate(conn, cmax + dt.timedelta(days=1), end, mode))

    combined = (pd.concat(parts, ignore_index=True)
                  .drop_duplicates(subset=["AccountNumber", "UnitKey", "Date"], keep="last"))
    if len(parts) > 1:
        combined.to_parquet(path, index=False)
    return _slice(combined, start, end)


# ── Features (doc Appendix A.1 + B.1, in pandas) ─────────────────────
def compute_player_features(daily: pd.DataFrame, as_of_date) -> pd.DataFrame:
    as_of = pd.to_datetime(as_of_date)
    cur_start = as_of - pd.Timedelta(days=30)
    prior_start = as_of - pd.Timedelta(days=60)
    d = daily[daily["Date"] <= as_of].copy()

    g = d.groupby("AccountNumber")
    feat = pd.DataFrame({
        "LCI": g["Handle"].sum(),
        "TotalSpins": g["Spins"].sum(),
        "LAD": g["Date"].nunique(),
        "MaxDate": g["Date"].max(),
    })
    cur = d[d["Date"] > cur_start].groupby("AccountNumber")
    feat["CI30"] = cur["Handle"].sum()
    feat["AD30"] = cur["Date"].nunique()
    pri = d[(d["Date"] > prior_start) & (d["Date"] <= cur_start)].groupby("AccountNumber")
    feat["CI30Prior"] = pri["Handle"].sum()

    feat[["CI30", "CI30Prior"]] = feat[["CI30", "CI30Prior"]].fillna(0.0)
    feat["AD30"] = feat["AD30"].fillna(0).astype(int)
    feat["Recency"] = (as_of - feat["MaxDate"]).dt.days
    feat = feat[(feat["LAD"] >= 3) & (feat["LCI"] > 0)].copy()
    feat.index.name = "AccountNumber"
    return feat.reset_index()


def engineer_features(df: pd.DataFrame) -> pd.DataFrame:
    X = pd.DataFrame(index=df.index)
    X["log_LCI"] = np.log1p(df["LCI"])
    X["log_CI30"] = np.log1p(df["CI30"])
    max_rec = df["Recency"].max()
    X["recency_score"] = 1.0 - (df["Recency"] / max_rec) if max_rec else 0.0
    X["recent_active_pct"] = df["AD30"] / 30.0
    X["log_LAD"] = np.log1p(df["LAD"])
    avg_bet = np.where(df["TotalSpins"] > 0, df["LCI"] / df["TotalSpins"], 0)
    X["log_avg_bet"] = np.log1p(np.minimum(avg_bet, 100))
    spins_day = np.where(df["LAD"] > 0, df["TotalSpins"] / df["LAD"], 0)
    X["log_spins_day"] = np.log1p(spins_day)
    trend = np.where(df["CI30Prior"] > 0, (df["CI30"] / df["CI30Prior"]) - 1.0,
                     np.where(df["CI30"] > 0, 5.0, 0.0))
    X["trend_clipped"] = np.clip(trend, -1, 5)
    return X[FEATURE_COLS]


def run_clustering(X, k: int = 6, seed: int = 42):
    scaler = StandardScaler()
    Xs = scaler.fit_transform(X.values)
    if len(Xs) > 100000:
        km = MiniBatchKMeans(n_clusters=k, n_init=10, random_state=seed,
                             batch_size=4096, max_iter=200)
    else:
        km = KMeans(n_clusters=k, n_init=25, random_state=seed, max_iter=500)
    return km.fit_predict(Xs), km.cluster_centers_, scaler


def name_clusters(centers: np.ndarray) -> dict:
    idx = {f: i for i, f in enumerate(FEATURE_COLS)}
    remaining = set(range(centers.shape[0]))
    names = {}

    def take(metric, how):
        col = idx[metric]
        c = sorted(remaining, key=lambda i: centers[i, col], reverse=(how == "max"))[0]
        remaining.discard(c)
        return c

    names[take("log_avg_bet", "max")] = "High-Roller"
    names[take("recent_active_pct", "max")] = "VIP"
    names[take("trend_clipped", "max")] = "Surging Spender"
    names[take("log_LCI", "min")] = "Light Tried"
    names[take("recency_score", "min")] = "Lapsed"
    names[remaining.pop()] = "Steady Regular"
    return names


def segment_from_daily(daily, locations, as_of, mode: str):
    feat = compute_player_features(daily, as_of)
    X = engineer_features(feat)
    labels, centers, _ = run_clustering(X)
    names = name_clusters(centers)
    feat["Cluster"] = pd.Series(labels, index=feat.index).map(names)

    sv = (daily.groupby(["AccountNumber", "UnitKey"])["Handle"].sum().reset_index()
                .sort_values(["AccountNumber", "Handle"], ascending=[True, False]))
    prim = sv.drop_duplicates("AccountNumber")[["AccountNumber", "UnitKey"]]
    if MODES[mode]["dim"] == "location":
        loc = locations.rename(columns={"LocationId": "UnitKey", "StateProv": "PrimaryState",
                                         "BusinessName": "PrimaryUnitName"})
        prim = prim.merge(loc[["UnitKey", "PrimaryState", "PrimaryUnitName"]],
                          on="UnitKey", how="left")
    else:
        prim["PrimaryState"] = np.nan
        prim["PrimaryUnitName"] = prim["UnitKey"]
    prim = prim.rename(columns={"UnitKey": "PrimaryUnitKey"})

    out = feat.merge(prim, on="AccountNumber", how="left")
    out["PerSpinBet"] = np.where(out["TotalSpins"] > 0, out["LCI"] / out["TotalSpins"], 0.0)
    out["AsOfDate"] = pd.to_datetime(as_of)

    cent = pd.DataFrame(centers, columns=FEATURE_COLS)
    cent.insert(0, "Cluster", [names[i] for i in range(centers.shape[0])])
    cent = cent.set_index("Cluster").reindex([c for c in CLUSTER_ORDER if c in names.values()])
    return out, cent


def cluster_summary(players: pd.DataFrame) -> pd.DataFrame:
    g = players.groupby("Cluster")
    s = pd.DataFrame({
        "Players": g.size(), "AvgBetVolume": g["LCI"].mean(),
        "MedianBetVolume": g["LCI"].median(), "AvgSpins": g["TotalSpins"].mean(),
        "MedianSpins": g["TotalSpins"].median(), "MedianVisitFreq": g["AD30"].median(),
        "MedianRecency": g["Recency"].median(), "SumLCI": g["LCI"].sum(),
    })
    s["PerSpinBet"] = np.where(s["AvgSpins"] > 0, s["AvgBetVolume"] / s["AvgSpins"], 0.0)
    s["PctPop"] = 100 * s["Players"] / s["Players"].sum()
    s["PctRevenue"] = 100 * s["SumLCI"] / s["SumLCI"].sum()
    return s.reindex([c for c in CLUSTER_ORDER if c in s.index])


# ── Games + bet levels ───────────────────────────────────────────────
def _casino_filter(casino):
    return f" AND CasinoName='{_esc(casino)}'" if casino and casino != "(All)" else ""


def _currency_filter(currency):
    return f" AND CurrencyName='{_esc(currency)}'" if currency and currency != "(All)" else ""


def load_game_summary(conn, start, end, mode, extra="") -> pd.DataFrame:
    sql = f"""
        SELECT GameId, SUM(Spins) AS Spins,
               SUM({HANDLE_SQL}) AS Bet, SUM({WIN_SQL}) AS Win,
               COUNT(DISTINCT AccountNumber) AS Players
        FROM BetSpinSummaryCashView3
        WHERE {_scope_where(mode)}{extra}
          AND "Date" >= '{start}' AND "Date" <= '{end}'
        GROUP BY GameId"""
    df = query_df(conn, sql)
    if df.empty:
        return df
    df["GameId"] = df["GameId"].astype(str).str.strip()
    for c in ["Spins", "Players"]:
        df[c] = pd.to_numeric(df[c]).astype("int64")
    df["Bet"] = pd.to_numeric(df["Bet"]).astype(float)
    df["Win"] = pd.to_numeric(df["Win"]).astype(float)
    df["RTP"] = np.where(df["Bet"] > 0, 100.0 * df["Win"] / df["Bet"], 0.0)
    df["Hold"] = 100.0 - df["RTP"]
    return df


def load_game_daily(conn, start, end, mode, extra="") -> pd.DataFrame:
    sql = f"""
        SELECT GameId, "Date", SUM({HANDLE_SQL}) AS Bet, SUM({WIN_SQL}) AS Win, SUM(Spins) AS Spins
        FROM BetSpinSummaryCashView3
        WHERE {_scope_where(mode)}{extra}
          AND "Date" >= '{start}' AND "Date" <= '{end}'
        GROUP BY GameId, "Date" """
    df = query_df(conn, sql)
    if df.empty:
        return df
    df["GameId"] = df["GameId"].astype(str).str.strip()
    df["Date"] = pd.to_datetime(df["Date"])
    df["Bet"] = pd.to_numeric(df["Bet"]).astype(float)
    df["Win"] = pd.to_numeric(df["Win"]).astype(float)
    df["Spins"] = pd.to_numeric(df["Spins"]).astype("int64")
    return df


def load_bet_level_summary(conn, start, end, mode, casino=None, currency=None, extra="") -> pd.DataFrame:
    sql = f"""
        SELECT GameId, Bet,
               SUM({HANDLE_SQL}) AS Handle, SUM({WIN_SQL}) AS Win,
               SUM(Spins) AS Spins, COUNT(DISTINCT AccountNumber) AS Players
        FROM BetSpinSummaryCashView3
        WHERE {_scope_where(mode)}{_casino_filter(casino)}{_currency_filter(currency)}{extra}
          AND "Date" >= '{start}' AND "Date" <= '{end}'
        GROUP BY GameId, Bet"""
    df = query_df(conn, sql)
    if df.empty:
        return df
    df["GameId"] = df["GameId"].astype(str).str.strip()
    df["Bet"] = pd.to_numeric(df["Bet"]).astype("int64")
    df["Handle"] = pd.to_numeric(df["Handle"]).astype(float)
    df["Win"] = pd.to_numeric(df["Win"]).astype(float)
    df["Spins"] = pd.to_numeric(df["Spins"]).astype("int64")
    df["Players"] = pd.to_numeric(df["Players"]).astype("int64")
    df["BetUSD"] = df["Bet"] / 100.0
    df["RTP"] = np.where(df["Handle"] > 0, 100.0 * df["Win"] / df["Handle"], 0.0)
    return df


def load_player_bet_detail(conn, account, start, end, mode) -> pd.DataFrame:
    sql = f"""
        SELECT GameId, Bet, SUM({HANDLE_SQL}) AS Handle, SUM({WIN_SQL}) AS Win, SUM(Spins) AS Spins
        FROM BetSpinSummaryCashView3
        WHERE {_scope_where(mode)} AND AccountNumber='{_esc(account)}'
          AND "Date" >= '{start}' AND "Date" <= '{end}'
        GROUP BY GameId, Bet"""
    df = query_df(conn, sql)
    if df.empty:
        return df
    df["GameId"] = df["GameId"].astype(str).str.strip()
    df["Bet"] = pd.to_numeric(df["Bet"]).astype("int64")
    df["Handle"] = pd.to_numeric(df["Handle"]).astype(float)
    df["Win"] = pd.to_numeric(df["Win"]).astype(float)
    df["Spins"] = pd.to_numeric(df["Spins"]).astype("int64")
    df["BetUSD"] = df["Bet"] / 100.0
    df["RTP"] = np.where(df["Handle"] > 0, 100.0 * df["Win"] / df["Handle"], 0.0)
    return df


def flag_hot_cold(bl: pd.DataFrame, min_spins: int = 500, threshold_pp: float = 3.0) -> pd.DataFrame:
    bl = bl.copy()
    tot = bl.groupby("GameId").agg(GH=("Handle", "sum"), GW=("Win", "sum"))
    game_rtp = pd.Series(np.where(tot["GH"] > 0, 100.0 * tot["GW"] / tot["GH"], 0.0),
                         index=tot.index)
    bl["GameRTP"] = bl["GameId"].map(game_rtp)
    bl["DeltaPP"] = bl["RTP"] - bl["GameRTP"]

    def lab(r):
        if r["Spins"] < min_spins:
            return "Low volume"
        if r["DeltaPP"] >= threshold_pp:
            return "\U0001F525 HOT"
        if r["DeltaPP"] <= -threshold_pp:
            return "❄️ COLD"
        return "Normal"

    bl["Flag"] = bl.apply(lab, axis=1)
    return bl


# ── Dimension helpers ────────────────────────────────────────────────
def load_casinos(conn, mode, start, end) -> pd.DataFrame:
    df = query_df(conn, f"""
        SELECT CasinoName, COUNT(DISTINCT AccountNumber) Players, SUM(Spins) Spins
        FROM BetSpinSummaryCashView3
        WHERE {_scope_where(mode)} AND "Date" >= '{start}' AND "Date" <= '{end}'
        GROUP BY CasinoName ORDER BY Players DESC""")
    if not df.empty:
        df["Players"] = pd.to_numeric(df["Players"]).astype("int64")
    return df


def load_currencies(conn, mode, start, end, casino=None) -> list:
    df = query_df(conn, f"""
        SELECT DISTINCT CurrencyName FROM BetSpinSummaryCashView3
        WHERE {_scope_where(mode)}{_casino_filter(casino)}
          AND "Date" >= '{start}' AND "Date" <= '{end}' AND CurrencyName IS NOT NULL""")
    return sorted([c for c in df["CurrencyName"].astype(str) if c and c != "None"])


# ── Game catalog ─────────────────────────────────────────────────────
def load_game_catalog_sql(conn) -> pd.DataFrame:
    sql = """
    SELECT Id, Name, Type, Platform, Product, Codebase, Status,
           MinBet, ScreenOrientation, JackpotStatus, Vip,
           Seasonal, Mechanics, Theme, Branded, SkinOf, ModifiedAt
    FROM GameCatalogView1
    """
    df = query_df(conn, sql)
    if df.empty:
        return df
    df["Id"] = pd.to_numeric(df["Id"], errors="coerce")
    df["Name"] = df["Name"].astype(str).str.strip()
    df["Codebase"] = df["Codebase"].astype(str).str.strip()
    return df


def load_game_catalog(path: str | None = None) -> pd.DataFrame:
    """Excel fallback -- kept only so load_game_catalog_with_fallback() has the
    same failure-mode shape as the original; the synthetic clone never ships a
    game_catalog.xlsx, so this simply raises and the caller returns an empty df."""
    import openpyxl
    path = path or os.path.join(DATA_DIR, "game_catalog.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["List"] if "List" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(rows[0])]
    cat = pd.DataFrame(rows[1:], columns=header)
    cat["KEY"] = cat["KEY"].astype(str).str.strip()
    return cat


def load_game_catalog_with_fallback(conn=None) -> pd.DataFrame:
    if conn is not None:
        try:
            df = load_game_catalog_sql(conn)
            if not df.empty:
                return df
        except Exception:
            pass
    try:
        return load_game_catalog()
    except Exception:
        return pd.DataFrame()


def attach_game_names(per_game: pd.DataFrame, catalog: pd.DataFrame,
                      prefixes=GAME_PREFIXES) -> pd.DataFrame:
    out = per_game.copy()
    out["GameId"] = out["GameId"].astype(str)
    if catalog is None or not len(catalog):
        out["GameName"] = prefixes[0] + " - " + out["GameId"]
        return out

    if "Id" in catalog.columns and "Name" in catalog.columns:
        lut = catalog.dropna(subset=["Name"]).drop_duplicates("Id").copy()
        lut["Id"] = pd.to_numeric(lut["Id"], errors="coerce")
        id_map = lut.set_index("Id")

        def _sql_lookup(gid):
            try:
                iid = int(gid)
                return id_map.at[iid, "Name"] if iid in id_map.index else None
            except Exception:
                return None

        out["GameName"] = out["GameId"].map(_sql_lookup)
        for col, src in [("GameType", "Type"), ("Product", "Product"),
                         ("CodeBase", "Codebase"), ("Platform", "Platform"),
                         ("Status", "Status")]:
            if src in id_map.columns:
                out[col] = out["GameId"].map(
                    lambda g, s=src: (id_map.at[int(g), s]
                                      if g.isdigit() and int(g) in id_map.index else None))
    else:
        cat = catalog.copy()
        cat["k"] = cat["KEY"].astype(str).str.replace(" ", "", regex=False)
        lut = cat.dropna(subset=["GAME NAME"]).drop_duplicates("k").set_index("k")

        def lookup(gid, col):
            for p in prefixes:
                key = f"{p}-{gid}"
                if key in lut.index:
                    return lut.at[key, col]
            return None

        out["GameName"] = out["GameId"].map(lambda g: lookup(g, "GAME NAME"))
        for col, src in [("GameType", "GAME TYPE"), ("Volatility", "VOLATILITY"),
                         ("Product", "PRODUCT"), ("CodeBase", "CODE BASE")]:
            if src in lut.columns:
                out[col] = out["GameId"].map(lambda g: lookup(g, src))

    out["Key"] = prefixes[0] + " - " + out["GameId"]
    out["GameName"] = out["GameName"].fillna(out["Key"])
    return out


def load_terminal_breakdown(conn, game_id: int, platform: str) -> pd.DataFrame:
    """Not called by launch.py/launch_dashboard_v2.py in this clone (verified by
    grep), and there is no AnalyticsGameTerminals table in the synthetic DB since
    nothing references it -- kept only as a stub so a stray import doesn't crash."""
    return pd.DataFrame()
